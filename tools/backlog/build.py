#!/usr/bin/env python3
"""Build JKX-Backlog.xlsx from the tab-separated tables in docs/backlog.

WHY THIS LIVES IN THE REPOSITORY. The first version of this generator lived in
a scratch directory beside the checkout, and a sandbox reset took it with
everything else - leaving the spreadsheet it produces, on somebody's disk, as
the only copy of a hundred and ninety rows that could no longer be regenerated.
A tool that builds a deliverable belongs next to the thing it builds.

WHY THE DATA IS NOT IN THIS FILE. The rows are in Russian and this is code:
rule 1 of the coding standards says code and comments are Latin-only, and the
gate that enforces it reads .py. So the tables sit in docs/backlog/*.tsv, which
is documentation by both the standard's definition and this file's - they are
edited by hand and read by a person. That split is not a workaround; it is the
right shape. The data outlives the presentation.

    tools/backlog/build.py [output.xlsx]

Columns are taken from the first row of each file, so adding one means editing
the .tsv and nothing else.
"""

import csv
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DATA = os.path.join(ROOT, "docs", "backlog")

BASE = Font(name="Calibri", size=11)
BOLD = Font(name="Calibri", size=11, bold=True)
TITLE = Font(name="Calibri", size=16, bold=True)
HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="44546A")
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(vertical="top", wrap_text=True)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read(name):
    """A tab-separated table from docs/backlog: (header row, data rows)."""
    path = os.path.join(DATA, name)
    with open(path, encoding="utf-8", newline="") as handle:
        rows = [r for r in csv.reader(handle, delimiter="\t") if any(r)]
    if not rows:
        raise SystemExit(f"{path}: no rows")
    return rows[0], rows[1:]


def read_pairs(name):
    """Two columns of a table as a dict, in file order."""
    _, rows = read(name)
    return {r[0]: r[1] for r in rows}


# Colours and every piece of text the spreadsheet shows come from docs/backlog
# for the same reason the rows do: they are Russian, and this file is code.
SIZE_FILL = {k: PatternFill("solid", fgColor=v)
             for k, v in read_pairs("sizes.tsv").items()}
STATE_FILL = {k: PatternFill("solid", fgColor=v)
              for k, v in read_pairs("states.tsv").items()}
LABEL = read_pairs("labels.tsv")

SIZES = tuple(SIZE_FILL)

# The order the summary lists states in is the order states.tsv lists them, and
# the reference row is not a state a task can be in - it colours cells that are
# there for context rather than for work.
STATES = tuple(k for k in STATE_FILL if k != LABEL["state.reference"])
OPEN = LABEL["state.open"]
TASKS = LABEL["sheet.tasks"]


def sheet(book, title, headers, rows, widths, freeze="A2"):
    ws = book.create_sheet(title)
    ws.sheet_view.showGridLines = False

    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    for r, row in enumerate(rows, start=2):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.font = BASE
            cell.alignment = WRAP
            cell.border = BORDER

    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    return ws


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, "JKX-Backlog.xlsx")

    task_head, tasks = read("tasks.tsv")
    batch_head, batch = read("batch1.tsv")
    find_head, findings = read("findings.tsv")

    unknown = sorted({r[5] for r in tasks if len(r) > 5 and r[5] not in STATE_FILL})
    if unknown:
        raise SystemExit("unknown state(s) in tasks.tsv: " + ", ".join(unknown))

    book = Workbook()
    book.remove(book.active)

    summary = book.create_sheet(LABEL["sheet.summary"])
    summary.sheet_view.showGridLines = False
    summary.column_dimensions["A"].width = 46
    summary.column_dimensions["B"].width = 12

    ws = sheet(book, LABEL["sheet.tasks"], task_head, tasks,
               [10, 8, 16, 62, 8, 14, 14, 16, 70])
    last = len(tasks) + 1

    for r in range(2, last + 1):
        size = ws.cell(row=r, column=5)
        size.alignment = CENTER
        if size.value in SIZE_FILL:
            size.fill = SIZE_FILL[size.value]
        state = ws.cell(row=r, column=6)
        state.alignment = CENTER
        state.fill = STATE_FILL.get(state.value, STATE_FILL[LABEL["state.reference"]])

    sheet(book, LABEL["sheet.batch"], batch_head, batch, [6, 10, 52, 30, 16, 52])
    sheet(book, LABEL["sheet.findings"], find_head, findings, [6, 62, 14, 22, 8, 22])

    summary.cell(row=1, column=1, value=LABEL["book.title"]).font = TITLE
    summary.cell(row=2, column=1, value=LABEL["book.subtitle"]).font = BASE

    # Counted by the spreadsheet rather than by this script on purpose: the
    # numbers then follow the filters a person applies, instead of describing
    # the moment the file was written.
    row = 4
    for label, size, state in (
            (LABEL["head.by_state"], None, None),
            *[(s, None, s) for s in STATES],
            ("", None, None),
            (LABEL["head.by_size"], None, None),
            *[(s, s, None) for s in SIZES],
            ("", None, None),
            (LABEL["head.open_by_size"], None, None),
            *[(s, s, OPEN) for s in SIZES]):
        cell = summary.cell(row=row, column=1, value=label)
        if size is None and state is None:
            cell.font = BOLD if label else BASE
        else:
            cell.font = BASE
            if size and state:
                f = (f'=COUNTIFS({TASKS}!$E$2:$E${last},"{size}",'
                     f'{TASKS}!$F$2:$F${last},"{state}")')
            elif size:
                f = f'=COUNTIF({TASKS}!$E$2:$E${last},"{size}")'
            else:
                f = f'=COUNTIF({TASKS}!$F$2:$F${last},"{state}")'
            value = summary.cell(row=row, column=2, value=f)
            value.font = BOLD
            value.alignment = CENTER
            if size:
                value.fill = SIZE_FILL[size]
        row += 1

    summary.cell(row=row + 1, column=1, value=LABEL["total"]).font = BOLD
    total = summary.cell(row=row + 1, column=2,
                         value=f'=COUNTA({TASKS}!$A$2:$A${last})')
    total.font = BOLD
    total.alignment = CENTER

    book.save(out)
    print("%s: %d task(s), %d in the small batch, %d finding(s)"
          % (out, len(tasks), len(batch), len(findings)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
